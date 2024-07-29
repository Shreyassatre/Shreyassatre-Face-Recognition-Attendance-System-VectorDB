from fastapi import UploadFile, HTTPException
from pinecone import Pinecone
import io
import cv2
from deepface import DeepFace
import numpy as np
from PIL import Image
import pandas as pd
from ..database import SessionLocal, User
from datetime import datetime
import os
from sqlalchemy.exc import NoResultFound
from openpyxl import Workbook, load_workbook

pc = Pinecone(api_key='5fae31d3-2cf4-474a-ad33-11826e76bf01')
index = pc.Index("face-recognition")

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

ATTENDANCE_DIR = "attendance_sheets"
os.makedirs(ATTENDANCE_DIR, exist_ok=True)

def get_embedding(image):
    image_np = np.array(image)
    
    embedding = DeepFace.represent(img_path=image_np, model_name='Facenet512')[0]['embedding']
    embedding = np.array(embedding, dtype=np.float32)
    if np.isnan(embedding).any() or np.isinf(embedding).any():
        raise ValueError("Embedding contains NaN or Inf values.")
    return embedding

def detect_face(image):
    image_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    
    gray = cv2.cvtColor(image_cv, cv2.COLOR_BGR2GRAY)
    
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(60, 60))
    
    if len(faces) == 0:
        raise ValueError("No faces detected in the image.")
    
    (x, y, w, h) = faces[0]
    
    face = image_cv[y:y+h, x:x+w]

    cv2.imwrite('processed.png', face)

    face_image = Image.fromarray(cv2.cvtColor(face, cv2.COLOR_BGR2RGB))
    
    return face_image

def check_existing_user(new_embedding, similarity_threshold):
    results = index.query(vector=new_embedding.tolist(), top_k=10)
    
    for match in results.get('matches', []):
        similarity_score = match['score']
        if similarity_score >= similarity_threshold:
            return True, similarity_score
    
    return False, None

def register_user(username, image, ip_address, similarity_threshold=0.70):
    db = SessionLocal()
    try:
        face_image = detect_face(image)
        
        new_embedding = get_embedding(face_image)
        
        exists, similarity_score = check_existing_user(new_embedding, similarity_threshold)
        
        if exists:
            return False, f"User with similar face already registered. Similarity score: {similarity_score:.2f}"
        
        new_user = User(
            username=username,
            ip=ip_address,
            timestamp=datetime.utcnow()
        )
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        
        index.upsert(vectors=[(str(new_user.id), new_embedding.tolist())])
        
        return True, f"User {username} registered successfully with ID {new_user.id}"
    except Exception as e:
        return False, str(e)
    finally:
        db.close()

def verify_user(image, confidence_threshold=0.70):
    try:
        face_image = detect_face(image)
        
        embedding = get_embedding(face_image)
        
        results = index.query(vector=embedding.tolist(), top_k=1)
        
        if results and results['matches']:
            match = results['matches'][0]
            user_id = match['id']
            similarity_score = match['score']
            
            if similarity_score >= confidence_threshold:
                return True, user_id, similarity_score
            else:
                return False, None, similarity_score
        else:
            return False, None, None
    except Exception as e:
        raise ValueError(f"Error during verification: {str(e)}")
    
async def read_image_data(upload_file: UploadFile):
    try:
        contents = await upload_file.read()
        image = Image.open(io.BytesIO(contents))
        return np.array(image)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

def get_username_by_id(user_id: int):
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).one()
        return user.username
    except NoResultFound:
        return None
    finally:
        db.close()

def log_attendance(user_id: str, username: str):
    today_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    file_path = os.path.join(ATTENDANCE_DIR, f"{today_date}.xlsx")
    
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Time", "User ID", "Username"])
    
    # for row in sheet.iter_rows(min_row=2, values_only=True):
    #     if row[2] == user_id and row[0] == today_date:
    #         return

    sheet.append([today_date, current_time, user_id, username])
    workbook.save(file_path)